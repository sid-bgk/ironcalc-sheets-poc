"""
Verify the newly added table named ranges
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX_PATH = r"D:\code\work\iron_calc_poc\DSCR_NoArrayFormulas_Testing.xlsx"

def main():
    print("=" * 80)
    print("VERIFICATION OF NEW TABLE NAMED RANGES")
    print("=" * 80)

    wb = load_workbook(XLSX_PATH, data_only=False)
    ws = wb["API_Output"]

    # =========================================================
    # 1. RateStackTable (G5:AK32)
    # =========================================================
    print("\n" + "=" * 80)
    print("1. RateStackTable (G5:AK32) - 28 rows x 31 cols")
    print("=" * 80)

    print("\nRow 5 (Headers):")
    headers = []
    for col in range(7, 38):
        cell = ws.cell(row=5, column=col)
        headers.append(str(cell.value)[:15] if cell.value else "")
    print(f"  {headers[:5]} ... {headers[-3:]}")

    print("\nRow 6 (First data row) - checking for formulas:")
    for col in range(7, 12):
        cell = ws.cell(row=6, column=col)
        col_letter = get_column_letter(col)
        val = cell.value
        is_formula = str(val).startswith("=") if val else False
        print(f"  {col_letter}6: {'FORMULA' if is_formula else 'STATIC'} - {str(val)[:50]}")

    # Count formulas in data area
    formula_count = 0
    for row in range(6, 33):
        for col in range(7, 38):
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value).startswith("="):
                formula_count += 1
    print(f"\nFormulas in data area (G6:AK32): {formula_count} / {27*31} cells")
    print(f"Classification should be: OUTPUT")

    # =========================================================
    # 2. PriceAdjustmentTable (A7:C22)
    # =========================================================
    print("\n" + "=" * 80)
    print("2. PriceAdjustmentTable (A7:C22) - 16 rows x 3 cols")
    print("=" * 80)

    print("\nFull table dump:")
    print(f"{'Row':<5} {'A':<25} {'B':<25} {'C':<25}")
    print("-" * 80)

    formula_count = 0
    for row_num in range(7, 23):
        row_data = []
        for col in range(1, 4):
            cell = ws.cell(row=row_num, column=col)
            val = cell.value
            if val and str(val).startswith("="):
                formula_count += 1
                row_data.append("[FORMULA]")
            elif val:
                row_data.append(str(val)[:23])
            else:
                row_data.append("")
        print(f"{row_num:<5} {row_data[0]:<25} {row_data[1]:<25} {row_data[2]:<25}")

    print(f"\nFormulas in table: {formula_count} / {16*3} cells")
    print(f"Classification should be: {'OUTPUT' if formula_count > 0 else 'OUTPUT (TABLE always OUTPUT)'}")

    # =========================================================
    # 3. EligiblityFailureReason (A30:B92)
    # =========================================================
    print("\n" + "=" * 80)
    print("3. EligiblityFailureReason (A30:B92) - 63 rows x 2 cols")
    print("=" * 80)

    print("\nFirst 15 rows:")
    print(f"{'Row':<5} {'A':<35} {'B':<40}")
    print("-" * 80)

    formula_count = 0
    for row_num in range(30, 45):
        row_data = []
        for col in range(1, 3):
            cell = ws.cell(row=row_num, column=col)
            val = cell.value
            if val and str(val).startswith("="):
                formula_count += 1
                row_data.append(f"[F:{str(val)[:30]}]")
            elif val:
                row_data.append(str(val)[:38])
            else:
                row_data.append("")
        print(f"{row_num:<5} {row_data[0]:<35} {row_data[1]:<40}")

    # Count all formulas
    total_formulas = 0
    for row_num in range(30, 93):
        for col in range(1, 3):
            cell = ws.cell(row=row_num, column=col)
            if cell.value and str(cell.value).startswith("="):
                total_formulas += 1

    print(f"\n... (showing first 15 of 63 rows)")
    print(f"\nFormulas in table: {total_formulas} / {63*2} cells")
    print(f"Classification should be: OUTPUT (TABLE always OUTPUT)")

    # =========================================================
    # Summary
    # =========================================================
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print("""
    All 3 tables should be classified as OUTPUT because:
    1. TABLE type = always OUTPUT (per our strategy)
    2. Even if first cell is header (no formula), data cells have formulas

    Current script classifies based on first cell only - needs update.
    """)

    wb.close()

if __name__ == "__main__":
    main()
