"""
Deep exploration of Input sheet table areas
Focus on A7:C23 and A30:B92 to understand their structure
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX_PATH = r"D:\code\work\iron_calc_poc\DSCR_NoArrayFormulas_Testing.xlsx"

def main():
    print("=" * 70)
    print("INPUT SHEET - DEEP TABLE EXPLORATION")
    print("=" * 70)

    wb = load_workbook(XLSX_PATH, data_only=False)
    ws = wb["Input"]

    # =========================================================
    # Table 1: A7:C23 - Full dump
    # =========================================================
    print("\n" + "=" * 70)
    print("TABLE 1: A7:C23 (Full dump)")
    print("=" * 70)
    print(f"{'Row':<5} {'A':<25} {'B':<25} {'C':<25} {'D':<25}")
    print("-" * 105)

    for row_num in range(7, 24):
        row_data = []
        for col in range(1, 5):  # A to D
            cell = ws.cell(row=row_num, column=col)
            val = cell.value
            if val is None:
                row_data.append("")
            elif str(val).startswith("="):
                row_data.append(f"[F:{str(val)[:15]}]")
            else:
                row_data.append(str(val)[:23])
        print(f"{row_num:<5} {row_data[0]:<25} {row_data[1]:<25} {row_data[2]:<25} {row_data[3]:<25}")

    # =========================================================
    # Table 2: A30:B92 - Full dump (sample)
    # =========================================================
    print("\n" + "=" * 70)
    print("TABLE 2: A30:B92 (showing rows 30-50)")
    print("=" * 70)
    print(f"{'Row':<5} {'A':<30} {'B':<30} {'C':<30} {'D':<30}")
    print("-" * 125)

    for row_num in range(30, 51):
        row_data = []
        for col in range(1, 5):  # A to D
            cell = ws.cell(row=row_num, column=col)
            val = cell.value
            if val is None:
                row_data.append("")
            elif str(val).startswith("="):
                row_data.append(f"[F:{str(val)[:20]}]")
            else:
                row_data.append(str(val)[:28])
        print(f"{row_num:<5} {row_data[0]:<30} {row_data[1]:<30} {row_data[2]:<30} {row_data[3]:<30}")

    # =========================================================
    # Explore where actual INPUT values are (Column D?)
    # =========================================================
    print("\n" + "=" * 70)
    print("COLUMN D ANALYSIS (where input values likely are)")
    print("=" * 70)
    print(f"{'Row':<5} {'C (Label)':<35} {'D (Value)':<25} {'Has Formula?':<15}")
    print("-" * 80)

    for row_num in range(6, 25):
        c_cell = ws.cell(row=row_num, column=3)
        d_cell = ws.cell(row=row_num, column=4)

        c_val = str(c_cell.value)[:33] if c_cell.value else ""
        d_val = d_cell.value
        has_formula = "YES" if d_val and str(d_val).startswith("=") else "NO"
        d_display = str(d_val)[:23] if d_val else ""

        if c_val or d_val:  # Only show non-empty rows
            print(f"{row_num:<5} {c_val:<35} {d_display:<25} {has_formula:<15}")

    # =========================================================
    # Explore the Liabilities test area (C60:C63 and D65:G65)
    # =========================================================
    print("\n" + "=" * 70)
    print("LIABILITIES TEST AREA (existing named ranges)")
    print("=" * 70)

    print("\nLiabilitiesVertical (C60:C63):")
    print("-" * 40)
    for row_num in range(58, 72):
        row_data = []
        for col in range(2, 9):  # B to H
            cell = ws.cell(row=row_num, column=col)
            val = cell.value
            if val is None:
                row_data.append("")
            elif str(val).startswith("="):
                row_data.append(f"[F]")
            else:
                row_data.append(str(val)[:10])
        print(f"Row {row_num}: {row_data}")

    # =========================================================
    # Find all areas with actual data (non-empty, non-formula cells in col D)
    # =========================================================
    print("\n" + "=" * 70)
    print("INPUT VALUE CELLS IN COLUMN D (potential INPUT named ranges)")
    print("=" * 70)

    input_cells = []
    for row_num in range(1, 100):
        d_cell = ws.cell(row=row_num, column=4)
        c_cell = ws.cell(row=row_num, column=3)

        if d_cell.value and not str(d_cell.value).startswith("="):
            label = str(c_cell.value) if c_cell.value else f"Row {row_num}"
            input_cells.append({
                "row": row_num,
                "label": label[:30],
                "value": d_cell.value,
                "ref": f"D{row_num}"
            })

    print(f"\nFound {len(input_cells)} potential input cells in column D:\n")
    for cell in input_cells[:20]:
        print(f"  {cell['ref']:<6} = {str(cell['value']):<15} ({cell['label']})")
    if len(input_cells) > 20:
        print(f"  ... and {len(input_cells) - 20} more")

    wb.close()
    print("\n" + "=" * 70)
    print("EXPLORATION COMPLETE")
    print("=" * 70)

if __name__ == "__main__":
    main()
