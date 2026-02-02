"""
Detect Array Formulas in xlsx file
Scans all sheets for CSE array formulas (ArrayFormula objects)
"""

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
import sys

input_file = sys.argv[1] if len(sys.argv) > 1 else "../DSCR_NoArrayFormulas_DEV_CLEAN_GS_CONVERTED.xlsx"

print(f"\n{'='*70}")
print(f"ARRAY FORMULA DETECTION REPORT")
print(f"{'='*70}")
print(f"File: {input_file}\n")

# Load workbook
wb = openpyxl.load_workbook(input_file, data_only=False)

array_formulas = []
total_formulas = 0

print(f"Scanning {len(wb.sheetnames)} sheets...\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_arrays = []
    sheet_formulas = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                # Check for ArrayFormula object
                if isinstance(cell.value, ArrayFormula):
                    sheet_arrays.append({
                        'cell': cell.coordinate,
                        'formula': cell.value.text,
                        'ref': getattr(cell.value, 'ref', 'N/A')
                    })
                # Check for regular formula
                elif str(cell.value).startswith("="):
                    sheet_formulas += 1

    total_formulas += sheet_formulas

    if sheet_arrays:
        print(f"[!] {sheet_name}: {len(sheet_arrays)} ARRAY FORMULAS FOUND")
        for arr in sheet_arrays:
            print(f"    {arr['cell']}: {arr['formula'][:60]}...")
            if arr['ref'] != 'N/A':
                print(f"         Array ref: {arr['ref']}")
        array_formulas.extend([(sheet_name, a) for a in sheet_arrays])
    else:
        print(f"[OK] {sheet_name}: No array formulas ({sheet_formulas} regular formulas)")

print(f"\n{'='*70}")
print(f"SUMMARY")
print(f"{'='*70}")
print(f"Total sheets scanned: {len(wb.sheetnames)}")
print(f"Total regular formulas: {total_formulas}")
print(f"Total array formulas: {len(array_formulas)}")

if array_formulas:
    print(f"\n[WARNING] Array formulas detected - these may cause IronCalc issues!")
    print(f"\nArray formula locations:")
    for sheet_name, arr in array_formulas:
        print(f"  - {sheet_name}!{arr['cell']}")
else:
    print(f"\n[SUCCESS] No array formulas detected - file should be IronCalc compatible!")

wb.close()
