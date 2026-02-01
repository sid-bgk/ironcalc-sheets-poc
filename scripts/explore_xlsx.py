"""
Explore xlsx file structure for Story 2.1
Analyzes sheets, named ranges, and key table areas
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re

XLSX_PATH = r"D:\code\work\iron_calc_poc\DSCR_NoArrayFormulas_Testing.xlsx"

def parse_range_ref(ref):
    """Parse a cell reference like 'Sheet!$A$1:$B$10' into components"""
    ref = ref.replace("$", "")
    parts = ref.split("!")
    if len(parts) == 2:
        sheet_ref = parts[0].strip("'")
        cell_range = parts[1]
    else:
        sheet_ref = None
        cell_range = parts[0]

    if ":" in cell_range:
        start, end = cell_range.split(":")
        start_match = re.match(r"([A-Z]+)(\d+)", start)
        end_match = re.match(r"([A-Z]+)(\d+)", end)
        if start_match and end_match:
            return {
                "sheet": sheet_ref,
                "start_col": start_match.group(1),
                "start_row": int(start_match.group(2)),
                "end_col": end_match.group(1),
                "end_row": int(end_match.group(2)),
                "is_range": True
            }
    else:
        match = re.match(r"([A-Z]+)(\d+)", cell_range)
        if match:
            return {
                "sheet": sheet_ref,
                "start_col": match.group(1),
                "start_row": int(match.group(2)),
                "end_col": match.group(1),
                "end_row": int(match.group(2)),
                "is_range": False
            }
    return None

def get_range_type(parsed):
    """Determine if range is SINGLE, HORIZONTAL, VERTICAL, or TABLE"""
    if not parsed or not parsed.get("is_range"):
        return "SINGLE"

    start_col_num = column_index_from_string(parsed["start_col"])
    end_col_num = column_index_from_string(parsed["end_col"])

    row_count = parsed["end_row"] - parsed["start_row"] + 1
    col_count = end_col_num - start_col_num + 1

    if row_count == 1 and col_count > 1:
        return "HORIZONTAL"
    elif row_count > 1 and col_count == 1:
        return "VERTICAL"
    elif row_count > 1 and col_count > 1:
        return "TABLE"
    return "SINGLE"

def main():
    print("=" * 70)
    print("XLSX EXPLORATION REPORT")
    print("=" * 70)
    print(f"File: {XLSX_PATH}\n")

    # Load workbook with data_only=False to see formulas
    wb = load_workbook(XLSX_PATH, data_only=False)

    # 1. List all sheets
    print("=" * 70)
    print("1. SHEETS")
    print("=" * 70)
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        print(f"  [{i}] {sheet_name}")
        print(f"      Dimensions: {ws.dimensions}")
        print(f"      Max row: {ws.max_row}, Max col: {ws.max_column}")
        print()

    # 2. All defined names (named ranges) - using correct API
    print("=" * 70)
    print("2. DEFINED NAMES (Named Ranges)")
    print("=" * 70)

    # Iterate through defined names correctly
    for name in wb.defined_names:
        defined_name = wb.defined_names[name]
        print(f"  Name: {name}")
        print(f"    Value/Formula: {defined_name.value}")
        print(f"    Attr Text: {defined_name.attr_text}")

        # Parse the reference
        parsed = parse_range_ref(defined_name.value)
        if parsed:
            range_type = get_range_type(parsed)
            print(f"    Sheet: {parsed['sheet']}")
            print(f"    Range: {parsed['start_col']}{parsed['start_row']}:{parsed['end_col']}{parsed['end_row']}")

            if parsed["is_range"]:
                start_col_num = column_index_from_string(parsed["start_col"])
                end_col_num = column_index_from_string(parsed["end_col"])
                row_count = parsed["end_row"] - parsed["start_row"] + 1
                col_count = end_col_num - start_col_num + 1
                print(f"    Dimensions: {row_count} rows x {col_count} cols")

            print(f"    Type: {range_type}")

            # Check if it contains formulas (for INPUT/OUTPUT classification)
            try:
                ws = wb[parsed["sheet"]]
                cell = ws.cell(row=parsed["start_row"], column=column_index_from_string(parsed["start_col"]))
                has_formula = str(cell.value).startswith("=") if cell.value else False
                print(f"    Classification: {'OUTPUT (has formula)' if has_formula else 'INPUT (no formula)'}")
                if has_formula:
                    print(f"    Sample formula: {str(cell.value)[:60]}...")
            except Exception as e:
                print(f"    Could not classify: {e}")

        print()

    # 3. Analyze Input sheet
    print("=" * 70)
    print("3. INPUT SHEET ANALYSIS")
    print("=" * 70)
    if "Input" in wb.sheetnames:
        ws = wb["Input"]

        # Table 1: A7:C23
        print("\n  Table A7:C23:")
        print("  " + "-" * 50)
        for row in ws.iter_rows(min_row=7, max_row=min(15, 23), min_col=1, max_col=3):
            values = [str(cell.value)[:20] if cell.value else "" for cell in row]
            print(f"    {values}")
        if 23 > 15:
            print(f"    ... ({23-15} more rows)")

        # Check for formulas in A7:C23
        print("\n  Formula check in A7:C23:")
        formula_count = 0
        for row in ws.iter_rows(min_row=7, max_row=23, min_col=1, max_col=3):
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    formula_count += 1
        print(f"    Total cells with formulas: {formula_count}")

        # Table 2: A30:B92
        print("\n  Table A30:B92:")
        print("  " + "-" * 50)
        for row in ws.iter_rows(min_row=30, max_row=min(38, 92), min_col=1, max_col=2):
            values = [str(cell.value)[:35] if cell.value else "" for cell in row]
            print(f"    {values}")
        if 92 > 38:
            print(f"    ... ({92-38} more rows)")

        # Check for formulas in A30:B92
        print("\n  Formula check in A30:B92:")
        formula_count = 0
        for row in ws.iter_rows(min_row=30, max_row=92, min_col=1, max_col=2):
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    formula_count += 1
        print(f"    Total cells with formulas: {formula_count}")

    # 4. Analyze API_Output sheet
    print("\n" + "=" * 70)
    print("4. API_OUTPUT SHEET ANALYSIS")
    print("=" * 70)
    if "API_Output" in wb.sheetnames:
        ws = wb["API_Output"]

        # Table G5:AK32 (G=7, AK=37)
        print("\n  Table G5:AK32 (main output table):")
        print("  " + "-" * 50)

        # Dimensions
        print(f"  Rows: 5 to 32 = 28 rows")
        print(f"  Cols: G to AK = 31 columns (G=7, AK=37)")

        # Header row (row 5)
        print("\n  Row 5 (Headers) - First 10 columns:")
        headers = []
        for col in range(7, 17):  # G=7 to P=16
            cell = ws.cell(row=5, column=col)
            headers.append(str(cell.value)[:12] if cell.value else "")
        print(f"    {headers}")

        print("\n  Row 5 (Headers) - Last 10 columns:")
        headers_last = []
        for col in range(28, 38):  # AB=28 to AK=37
            cell = ws.cell(row=5, column=col)
            headers_last.append(str(cell.value)[:12] if cell.value else "")
        print(f"    {headers_last}")

        # Sample data rows
        print("\n  Sample data (rows 6-10, first 5 cols):")
        for row_num in range(6, 11):
            row_data = []
            for col in range(7, 12):  # G to K
                cell = ws.cell(row=row_num, column=col)
                val = cell.value
                if val is None:
                    row_data.append("None")
                elif str(val).startswith("="):
                    row_data.append(f"[FORMULA]")
                else:
                    row_data.append(str(val)[:10])
            print(f"    Row {row_num}: {row_data}")

        # Formula analysis in G5:AK32
        print("\n  Formula analysis in G5:AK32:")
        formula_cells = 0
        static_cells = 0
        empty_cells = 0

        for row in range(5, 33):
            for col in range(7, 38):
                cell = ws.cell(row=row, column=col)
                if cell.value is None or cell.value == "":
                    empty_cells += 1
                elif str(cell.value).startswith("="):
                    formula_cells += 1
                else:
                    static_cells += 1

        total = formula_cells + static_cells + empty_cells
        print(f"    Total cells: {total}")
        print(f"    Cells with formulas: {formula_cells}")
        print(f"    Cells with static values: {static_cells}")
        print(f"    Empty cells: {empty_cells}")

        # Sample formulas
        print("\n  Sample formulas from G5:AK32:")
        found = 0
        for row in range(5, 33):
            for col in range(7, 38):
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).startswith("=") and found < 3:
                    coord = f"{get_column_letter(col)}{row}"
                    print(f"    {coord}: {str(cell.value)[:70]}...")
                    found += 1
            if found >= 3:
                break

    # 5. Summary for named range strategy
    print("\n" + "=" * 70)
    print("5. SUMMARY - EXISTING NAMED RANGES")
    print("=" * 70)
    print("\n  Count by type:")
    type_counts = {"SINGLE": 0, "HORIZONTAL": 0, "VERTICAL": 0, "TABLE": 0}
    classification_counts = {"INPUT": 0, "OUTPUT": 0, "UNKNOWN": 0}

    for name in wb.defined_names:
        defined_name = wb.defined_names[name]
        parsed = parse_range_ref(defined_name.value)
        if parsed:
            range_type = get_range_type(parsed)
            type_counts[range_type] += 1

            try:
                ws = wb[parsed["sheet"]]
                cell = ws.cell(row=parsed["start_row"], column=column_index_from_string(parsed["start_col"]))
                has_formula = str(cell.value).startswith("=") if cell.value else False
                if has_formula:
                    classification_counts["OUTPUT"] += 1
                else:
                    classification_counts["INPUT"] += 1
            except:
                classification_counts["UNKNOWN"] += 1

    for range_type, count in type_counts.items():
        print(f"    {range_type}: {count}")

    print("\n  Count by classification:")
    for classification, count in classification_counts.items():
        print(f"    {classification}: {count}")

    wb.close()
    print("\n" + "=" * 70)
    print("EXPLORATION COMPLETE")
    print("=" * 70)

if __name__ == "__main__":
    main()
