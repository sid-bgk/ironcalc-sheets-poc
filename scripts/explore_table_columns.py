"""
Deep exploration of G5:AK32 table columns
Analyze headers and data to determine best named range strategy
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re

XLSX_PATH = r"D:\code\work\iron_calc_poc\DSCR_NoArrayFormulas_Testing.xlsx"

def to_valid_name(header):
    """Convert header to valid Excel named range name"""
    # Remove special chars, replace spaces with underscores
    name = re.sub(r'[^a-zA-Z0-9_]', '', header.replace(' ', '_'))
    # Ensure starts with letter or underscore
    if name and name[0].isdigit():
        name = '_' + name
    return name

def main():
    print("=" * 80)
    print("TABLE COLUMN ANALYSIS - G5:AK32")
    print("=" * 80)

    wb = load_workbook(XLSX_PATH, data_only=False)
    ws = wb["API_Output"]

    # =========================================================
    # Analyze all column headers (Row 5, G to AK)
    # =========================================================
    print("\n" + "=" * 80)
    print("COLUMN HEADERS (Row 5)")
    print("=" * 80)
    print(f"\n{'Col':<4} {'Letter':<6} {'Header':<30} {'Valid Name':<30} {'Data Type'}")
    print("-" * 90)

    columns = []
    for col_num in range(7, 38):  # G=7 to AK=37
        col_letter = get_column_letter(col_num)
        header_cell = ws.cell(row=5, column=col_num)
        header = str(header_cell.value) if header_cell.value else f"Column{col_letter}"

        # Sample data from row 6 to determine type
        data_cell = ws.cell(row=6, column=col_num)
        data_val = data_cell.value

        if data_val is None:
            data_type = "EMPTY"
        elif str(data_val).startswith("="):
            # It's a formula - need to check calculated value
            data_type = "FORMULA"
        elif isinstance(data_val, (int, float)):
            data_type = "NUMBER"
        elif isinstance(data_val, bool):
            data_type = "BOOLEAN"
        else:
            data_type = "TEXT"

        valid_name = to_valid_name(header)

        columns.append({
            "col_num": col_num,
            "col_letter": col_letter,
            "header": header,
            "valid_name": valid_name,
            "data_type": data_type,
            "range": f"{col_letter}6:{col_letter}32"
        })

        print(f"{col_num:<4} {col_letter:<6} {header:<30} {valid_name:<30} {data_type}")

    # =========================================================
    # Grouping analysis - do headers suggest natural groups?
    # =========================================================
    print("\n" + "=" * 80)
    print("HEADER GROUPING ANALYSIS")
    print("=" * 80)

    # Find common prefixes
    prefixes = {}
    for col in columns:
        header = col["header"]
        # Check for common prefixes
        if header.startswith("chk_"):
            prefix = "chk (checks)"
        elif header.startswith("is"):
            prefix = "is (booleans)"
        elif header.startswith("max"):
            prefix = "max (limits)"
        elif header.startswith("total"):
            prefix = "total (aggregates)"
        else:
            prefix = "(other)"

        if prefix not in prefixes:
            prefixes[prefix] = []
        prefixes[prefix].append(col["header"])

    print("\nHeaders by prefix pattern:")
    for prefix, headers in prefixes.items():
        print(f"\n  {prefix}: ({len(headers)} columns)")
        for h in headers:
            print(f"    - {h}")

    # =========================================================
    # Proposed naming convention
    # =========================================================
    print("\n" + "=" * 80)
    print("PROPOSED NAMED RANGES (Column-per-range approach)")
    print("=" * 80)

    print("\nOption A: Prefix with 'Pricing_'")
    print("-" * 60)
    for col in columns[:10]:
        name = f"Pricing_{col['valid_name']}"
        print(f"  {name:<45} → API_Output!${col['col_letter']}$6:${col['col_letter']}$32")
    print(f"  ... ({len(columns) - 10} more)")

    print("\nOption B: Prefix with 'OUT_' (to indicate OUTPUT)")
    print("-" * 60)
    for col in columns[:10]:
        name = f"OUT_{col['valid_name']}"
        print(f"  {name:<45} → API_Output!${col['col_letter']}$6:${col['col_letter']}$32")
    print(f"  ... ({len(columns) - 10} more)")

    print("\nOption C: No prefix (just use header name)")
    print("-" * 60)
    for col in columns[:10]:
        name = col['valid_name']
        print(f"  {name:<45} → API_Output!${col['col_letter']}$6:${col['col_letter']}$32")
    print(f"  ... ({len(columns) - 10} more)")

    # =========================================================
    # Check for naming conflicts with existing named ranges
    # =========================================================
    print("\n" + "=" * 80)
    print("NAMING CONFLICT CHECK")
    print("=" * 80)

    existing_names = set()
    for name in wb.defined_names:
        existing_names.add(name.lower())

    conflicts = []
    for col in columns:
        if col['valid_name'].lower() in existing_names:
            conflicts.append(col['valid_name'])

    if conflicts:
        print(f"\n  Conflicts found: {conflicts}")
    else:
        print(f"\n  No naming conflicts with existing {len(existing_names)} named ranges")

    # =========================================================
    # Sample data from first few columns
    # =========================================================
    print("\n" + "=" * 80)
    print("SAMPLE DATA (first 5 columns, first 5 rows)")
    print("=" * 80)

    # Load with data_only=True to see calculated values
    wb_data = load_workbook(XLSX_PATH, data_only=True)
    ws_data = wb_data["API_Output"]

    print(f"\n{'Row':<5}", end="")
    for col in columns[:5]:
        print(f"{col['header']:<15}", end="")
    print()
    print("-" * 80)

    for row_num in range(6, 12):
        print(f"{row_num:<5}", end="")
        for col in columns[:5]:
            cell = ws_data.cell(row=row_num, column=col['col_num'])
            val = cell.value
            if val is None:
                display = "None"
            elif isinstance(val, float):
                display = f"{val:.4f}"[:14]
            else:
                display = str(val)[:14]
            print(f"{display:<15}", end="")
        print()

    # =========================================================
    # Alternative: Single table range with metadata
    # =========================================================
    print("\n" + "=" * 80)
    print("ALTERNATIVE: Single Table Range")
    print("=" * 80)
    print(f"""
  Named Range: PricingTable → API_Output!$G$5:$AK$32

  The code would:
  1. Detect it's a TABLE (28 rows × 31 cols)
  2. Treat row 1 as headers automatically
  3. Return structured JSON:

  {{
    "PricingTable": {{
      "headers": ["rate", "basePrice", ...],
      "data": [
        [0.065, 100.0, ...],
        [0.07, 99.5, ...],
        ...
      ]
    }}
  }}
""")

    # =========================================================
    # My recommendation
    # =========================================================
    print("=" * 80)
    print("RECOMMENDATION")
    print("=" * 80)
    print("""
  HYBRID APPROACH:

  1. For OUTPUTS where you want the WHOLE table:
     → Single named range: PricingTable → G5:AK32
     → Code detects TABLE type and returns structured JSON with headers

  2. For OUTPUTS where you want SPECIFIC columns:
     → Individual named ranges: Pricing_rate → G6:G32
     → These are VERTICAL arrays, returned as 1D arrays

  3. For grouping related columns:
     → Use naming prefix: Pricing_rate, Pricing_basePrice, etc.
     → API can optionally group by prefix

  This gives maximum flexibility:
  - Get whole table: use PricingTable
  - Get single column: use Pricing_rate
  - Consumer chooses what they need
""")

    wb.close()
    wb_data.close()

if __name__ == "__main__":
    main()
