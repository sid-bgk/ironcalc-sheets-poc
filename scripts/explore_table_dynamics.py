"""
Explore dynamic table scenarios and edge cases
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

XLSX_PATH = r"D:\code\work\iron_calc_poc\DSCR_NoArrayFormulas_Testing.xlsx"

def main():
    print("=" * 80)
    print("DYNAMIC TABLE ANALYSIS - Edge Cases & Solutions")
    print("=" * 80)

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["API_Output"]

    # =========================================================
    # Check actual data lengths per column
    # =========================================================
    print("\n" + "=" * 80)
    print("1. ACTUAL DATA LENGTH PER COLUMN (G6:AK32 area)")
    print("=" * 80)
    print("\nChecking how many non-empty cells each column has...")
    print(f"\n{'Col':<4} {'Header':<25} {'Non-Empty':<12} {'Empty':<12} {'Last Row'}")
    print("-" * 70)

    # Get headers from row 5
    wb_formula = load_workbook(XLSX_PATH, data_only=False)
    ws_formula = wb_formula["API_Output"]

    column_stats = []
    for col_num in range(7, 38):  # G=7 to AK=37
        col_letter = get_column_letter(col_num)
        header = ws_formula.cell(row=5, column=col_num).value or f"Col{col_letter}"

        non_empty = 0
        last_row_with_data = 5
        empty_count = 0

        for row_num in range(6, 33):  # Data rows 6-32
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value is not None and cell.value != "":
                non_empty += 1
                last_row_with_data = row_num
            else:
                empty_count += 1

        column_stats.append({
            "col": col_letter,
            "header": str(header)[:24],
            "non_empty": non_empty,
            "empty": empty_count,
            "last_row": last_row_with_data
        })

        print(f"{col_letter:<4} {str(header)[:24]:<25} {non_empty:<12} {empty_count:<12} {last_row_with_data}")

    # Check if all columns have same length
    non_empty_counts = set(c["non_empty"] for c in column_stats)
    last_rows = set(c["last_row"] for c in column_stats)

    print(f"\nUnique non-empty counts: {non_empty_counts}")
    print(f"Unique last rows: {last_rows}")

    if len(non_empty_counts) == 1:
        print("\n*** ALL COLUMNS HAVE SAME DATA LENGTH - Table is consistent! ***")
    else:
        print("\n*** WARNING: Columns have DIFFERENT lengths! ***")

    # =========================================================
    # Solution Analysis
    # =========================================================
    print("\n" + "=" * 80)
    print("2. SOLUTION OPTIONS FOR DYNAMIC TABLES")
    print("=" * 80)

    print("""
    PROBLEM: Column-per-named-range loses table structure info

    SOLUTION OPTIONS:

    A) METADATA NAMED RANGE
       Create a special named range that describes tables:

       Name: _TableMeta_Pricing
       Value: "cols=rate,basePrice,dscr;rows=6:32;sheet=API_Output"

       Pros: Explicit, flexible
       Cons: Extra maintenance, parsing needed

    B) API INTROSPECTION ENDPOINT
       GET /api/v1/calculate/dscr/schema

       Returns:
       {
         "tables": {
           "Pricing": {
             "columns": ["rate", "basePrice", "dscr", ...],
             "rowRange": "6:32",
             "sheet": "API_Output"
           }
         },
         "inputs": [...],
         "outputs": [...]
       }

       Pros: Clean separation, discoverable
       Cons: Extra endpoint, may get out of sync

    C) RETURN METADATA WITH VALUES
       Each output includes its dimensions:

       {
         "outputs": {
           "Pricing_rate": {
             "type": "VERTICAL",
             "rows": 27,
             "startRow": 6,
             "endRow": 32,
             "group": "Pricing",
             "values": [0.065, 0.07, ...]
           }
         }
       }

       Pros: Self-describing, always in sync
       Cons: Verbose response

    D) SINGLE TABLE NAMED RANGE (Original approach)
       One named range for entire table:

       PricingTable -> API_Output!$G$5:$AK$32

       Response:
       {
         "outputs": {
           "PricingTable": {
             "headers": ["rate", "basePrice", ...],
             "data": [[0.065, 100, ...], ...]
           }
         }
       }

       Pros: Simple, guaranteed consistent
       Cons: All or nothing, less flexible

    E) HYBRID: Table + Optional Columns
       Define BOTH:
       - PricingTable -> G5:AK32 (whole table)
       - Pricing_rate -> G6:G32 (specific column, optional)

       Consumer chooses:
       - Want whole table? Use PricingTable
       - Want one column? Use Pricing_rate

       Pros: Maximum flexibility
       Cons: Redundancy, maintenance burden
    """)

    # =========================================================
    # My Recommendation
    # =========================================================
    print("\n" + "=" * 80)
    print("3. RECOMMENDED APPROACH")
    print("=" * 80)

    print("""
    BEST APPROACH: Option D (Single Table Named Range) + Smart Code

    Why:
    1. GUARANTEED consistency - all columns same length
    2. SIMPLE to maintain - one named range per table
    3. FLEXIBLE output - code can format response smartly
    4. DYNAMIC - add/remove columns by changing range, headers auto-discovered

    Implementation:

    Named Range:
      PricingTable -> API_Output!$G$5:$AK$32

    Detection Logic:
      - If range has rows > 1 AND cols > 1 -> TABLE
      - First row = headers (auto-extracted)
      - Remaining rows = data

    Response Options (consumer chooses via query param?):

    Option 1: Full table with headers
    GET /api/v1/calculate/dscr?format=table

    {
      "outputs": {
        "PricingTable": {
          "headers": ["rate", "basePrice", "dscr", ...],
          "data": [
            [0.065, 100, 1.25, ...],
            [0.07, 99.5, 1.30, ...]
          ]
        }
      }
    }

    Option 2: Array of objects (each row is an object)
    GET /api/v1/calculate/dscr?format=objects

    {
      "outputs": {
        "PricingTable": [
          {"rate": 0.065, "basePrice": 100, "dscr": 1.25, ...},
          {"rate": 0.07, "basePrice": 99.5, "dscr": 1.30, ...}
        ]
      }
    }

    Option 3: Column-oriented (for when you want specific columns)
    GET /api/v1/calculate/dscr?format=columns

    {
      "outputs": {
        "PricingTable": {
          "rate": [0.065, 0.07, ...],
          "basePrice": [100, 99.5, ...],
          "dscr": [1.25, 1.30, ...]
        }
      }
    }

    This way:
    - ONE named range to maintain
    - Consumer gets flexibility via format parameter
    - Headers always included/discoverable
    - No mismatch possible between column lengths
    """)

    wb.close()
    wb_formula.close()

if __name__ == "__main__":
    main()
