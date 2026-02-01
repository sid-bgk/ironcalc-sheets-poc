"""
Clean DSCR_NoArrayFormulas_DEV.xlsx:
1. Clear all input values in Column D of Input sheet (preserve formulas)
2. Add named ranges based on Column C labels
"""

import openpyxl
import re

def label_to_name(label):
    """Convert a label like 'FICO Score' to a valid named range like 'FICOScore'"""
    if not label:
        return None

    # Remove parentheses and their contents
    name = re.sub(r'\s*\([^)]*\)', '', label)

    # Remove special characters except letters, numbers, underscores
    name = re.sub(r'[^a-zA-Z0-9_\s]', '', name)

    # Convert to PascalCase (remove spaces, capitalize each word)
    words = name.split()
    name = ''.join(word.capitalize() if word.islower() else word for word in words)

    # Ensure it starts with a letter or underscore
    if name and not name[0].isalpha() and name[0] != '_':
        name = '_' + name

    return name if name else None


def main():
    input_file = 'DSCR_NoArrayFormulas_DEV.xlsx'
    output_file = 'DSCR_NoArrayFormulas_DEV_CLEAN.xlsx'

    print(f'Loading {input_file}...')
    wb = openpyxl.load_workbook(input_file)
    ws = wb['Input']

    # Track what we do
    cleared_cells = []
    named_ranges = []
    skipped_formulas = []

    # Process rows with data in Column C
    for row in range(1, 100):
        c_val = ws.cell(row=row, column=3).value  # Column C (label)
        d_cell = ws.cell(row=row, column=4)        # Column D (value)

        if c_val is None:
            continue

        # Convert label to valid name
        range_name = label_to_name(c_val)
        if not range_name:
            continue

        # Check if D cell has a formula
        d_val = d_cell.value
        is_formula = isinstance(d_val, str) and d_val.startswith('=')

        if is_formula:
            skipped_formulas.append((row, c_val, d_val))
        else:
            # Clear the value (set to None to preserve formatting)
            if d_val is not None:
                cleared_cells.append((row, c_val, d_val))
                d_cell.value = None

        # Create named range pointing to this cell
        # Reference format: 'Input'!$D$5
        ref = f"'Input'!$D${row}"

        try:
            # Check if name already exists
            if range_name in wb.defined_names:
                print(f'  Warning: {range_name} already exists, skipping')
                continue

            wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(range_name, attr_text=ref))
            named_ranges.append((range_name, ref, c_val))
        except Exception as e:
            print(f'  Error creating {range_name}: {e}')

    # Save the workbook
    print(f'\nSaving to {output_file}...')
    wb.save(output_file)

    # Print summary
    print(f'\n=== Summary ===')
    print(f'\nCleared {len(cleared_cells)} input values:')
    for row, label, old_val in cleared_cells:
        print(f'  Row {row}: {label} = {old_val} -> (cleared)')

    print(f'\nSkipped {len(skipped_formulas)} formulas:')
    for row, label, formula in skipped_formulas:
        print(f'  Row {row}: {label} = {formula}')

    print(f'\nCreated {len(named_ranges)} named ranges:')
    for name, ref, label in named_ranges:
        print(f'  {name} -> {ref} ({label})')

    print(f'\nDone! Output saved to: {output_file}')


if __name__ == '__main__':
    main()
